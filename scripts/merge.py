# -*- coding: utf-8 -*-
"""日报合并脚本：读取当天 md 需求记录 → 追加到当月 Excel
用法：python merge.py
"""
import openpyxl, os, re, shutil
from datetime import datetime, timedelta
from copy import copy
from openpyxl.styles import Alignment

# ===== 配置 =====
DESKTOP = r"C:\Users\Administrator\Desktop"
TODAY = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
YEAR = TODAY.strftime("%Y")
MONTH = str(TODAY.month)  # 不用 strftime 避免补零（7月不是07月）
DAY = str(TODAY.day)
MM = TODAY.strftime("%m")  # 文件名里 MM-DD 需要补零
DD = TODAY.strftime("%d")

# 仓库名 → Excel 项目名 映射
REPO_MAP = {
    'lanxum-amisp': '档案V6',
    'lanxum-amisp-java': '档案V6',
    'lanxum-amisp-react': '档案V6',
    'workingpaper-v5.5': '中信底稿v5',
}

REPORT_DIR = os.path.join(DESKTOP, f"报告-{YEAR}年", f"日报-{YEAR}-{MONTH}月")
MD_FILE = os.path.join(REPORT_DIR, f"日报需求记录-{YEAR}-{MM}-{DD}.md")
XLSX_FILE = os.path.join(REPORT_DIR, f"日报表格-胡志伟~~{MM}-{DD}.xlsx")
TEMPLATE_FILE = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'templates', '日报模板.xlsx')

# ===== 工具函数 =====
def next_workday(date):
    """返回下一个工作日（跳过周六日）"""
    d = date + timedelta(days=1)
    while d.weekday() >= 5:  # 5=Sat, 6=Sun
        d += timedelta(days=1)
    return d

def is_workday(date):
    return date.weekday() < 5

def parse_md_table(filepath):
    """解析 md 表格，提取已完成任务 [{仓库, 需求概述, 人工工时, AI工时, 备注}]"""
    if not os.path.exists(filepath):
        raise FileNotFoundError(f"md 文件不存在: {filepath}")

    with open(filepath, 'r', encoding='utf-8') as f:
        content = f.read()

    tasks = []
    in_table = False
    for line in content.split('\n'):
        line = line.strip()
        if line.startswith('| 序号 |'):
            in_table = True
            continue
        if line.startswith('|------|'):
            continue
        if not in_table or not line.startswith('|'):
            continue
        if line.startswith('| >') or '空行为模板' in line:
            break

        cells = [c.strip() for c in line.split('|')[1:-1]]
        if len(cells) < 8:
            continue

        seq, date, repo, desc, modules, status, human_h, ai_h, note = cells[:9]

        # 兼容旧格式"已完成"和新格式百分比
        is_done = status in ('已完成', '100%')
        if is_done and seq.isdigit():
            # 提取百分比数字，默认 100
            pct = status.replace('%', '') if '%' in status else ('100' if is_done else '0')
            tasks.append({
                'repo': repo,
                'desc': desc,
                'human_h': float(human_h) if human_h else 0,
                'ai_h': float(ai_h) if ai_h else 0,
                'note': note,
                'pct': f'{pct}%',
            })

    return tasks

def find_base_excel():
    """找基准 Excel：当天已存在则用它，否则找最近一天的"""
    if os.path.exists(XLSX_FILE):
        return XLSX_FILE, True  # 当天已存在

    # 往前找最近的文件
    d = TODAY - timedelta(days=1)
    while d >= datetime(TODAY.year, TODAY.month, 1):
        path = os.path.join(REPORT_DIR, f"日报表格-胡志伟~~{d.strftime('%m-%d')}.xlsx")
        if os.path.exists(path):
            return path, False
        d -= timedelta(days=1)
    # 月初没上个月文件，用空白模板
    if os.path.exists(TEMPLATE_FILE):
        return TEMPLATE_FILE, False
    return None, False

def get_last_task_info(ws):
    """获取最后一条数据行的 G 列日期和当天剩余小时"""
    last_g = TODAY
    remaining_hours = 0

    # 找最后一条有数据的行（在备注说明之前）
    last_row = 0
    for row in range(1, ws.max_row + 1):
        c_val = ws.cell(row=row, column=3).value
        a_val = ws.cell(row=row, column=1).value
        if a_val and '填充说明' in str(a_val):
            break
        if c_val is not None and str(c_val).strip().isdigit():
            last_row = row

    if last_row > 0:
        g_val = ws.cell(row=last_row, column=7).value  # G 列
        if g_val:
            if isinstance(g_val, datetime):
                last_g = g_val
            else:
                try:
                    last_g = datetime.strptime(str(g_val)[:10], '%Y-%m-%d')
                except:
                    pass

        # 计算该日期已累计的小时数（当天所有任务 H 列之和 % 8 的余数不可靠）
        # 简化处理：从 G 列所在日期的第一个任务开始累加
        day_hours = 0
        for row in range(last_row, 0, -1):
            row_g = ws.cell(row=row, column=7).value
            row_g_date = None
            if row_g:
                if isinstance(row_g, datetime):
                    row_g_date = row_g.date()
                else:
                    try:
                        row_g_date = datetime.strptime(str(row_g)[:10], '%Y-%m-%d').date()
                    except:
                        pass
            if row_g_date != last_g.date():
                break
            h_val = ws.cell(row=row, column=8).value
            if h_val:
                try:
                    day_hours += float(h_val)
                except:
                    pass

        remaining_hours = day_hours % 8

    return last_g, remaining_hours

def find_notes_row(ws):
    """找备注说明起始行"""
    for row in range(1, ws.max_row + 1):
        a_val = ws.cell(row=row, column=1).value
        if a_val and '填充说明' in str(a_val):
            return row
    return ws.max_row + 3

def calc_g_column(hours, prev_date, remaining):
    """计算 G 列：从 prev_date 开始，按 8h/工作日叠加"""
    total = remaining + hours
    result_date = prev_date

    # 先消耗当天剩余容量
    while total > 8:
        total -= 8
        result_date = next_workday(result_date)

    return result_date, total  # 返回日期和当天剩余小时

def copy_style(src_cell, dst_cell):
    """复制单元格样式（不复制底色，避免表头蓝底污染数据行）"""
    if src_cell.has_style:
        dst_cell.font = copy(src_cell.font)
        dst_cell.border = copy(src_cell.border)
        dst_cell.alignment = copy(src_cell.alignment)

def remove_today_rows(ws, today_date):
    """删除当天已存在的行（用于去重重跑），正确处理合并单元格"""
    # 先找 A 列哪些合并区域包含今天的日期
    today_merged_rows = set()
    for mr in list(ws.merged_cells.ranges):
        if mr.min_col == 1 and mr.max_col == 1:  # A 列合并
            top_val = ws.cell(row=mr.min_row, column=1).value
            if top_val:
                if isinstance(top_val, datetime):
                    ad = top_val.date()
                else:
                    try:
                        ad = datetime.strptime(str(top_val)[:10], '%Y-%m-%d').date()
                    except:
                        continue
                if ad == today_date.date():
                    for r in range(mr.min_row, mr.max_row + 1):
                        today_merged_rows.add(r)
                    ws.unmerge_cells(str(mr))  # 拆掉这个合并

    # 找非合并区域中今天的行
    rows_to_delete = set()
    notes = find_notes_row(ws)
    for row in range(1, notes):
        a_val = ws.cell(row=row, column=1).value
        if a_val:
            if isinstance(a_val, datetime):
                ad = a_val.date()
            else:
                try:
                    ad = datetime.strptime(str(a_val)[:10], '%Y-%m-%d').date()
                except:
                    continue
            if ad == today_date.date():
                rows_to_delete.add(row)

    # 合并两个集合
    rows_to_delete |= today_merged_rows

    # 从下往上删，避免行号变化
    for row in sorted(rows_to_delete, reverse=True):
        ws.delete_rows(row)
        print(f"  删除已有行: {row}")

    return len(rows_to_delete)

# ===== 主流程 =====
def main():
    # 1. 读 md
    tasks = parse_md_table(MD_FILE)
    if not tasks:
        print("没有已完成的任务，跳过合并。")
        return

    print(f"找到 {len(tasks)} 条已完成任务：")
    for t in tasks:
        print(f"  - [{t['repo']}] {t['desc'][:50]}... ({t['human_h']}h + AI {t['ai_h']}h)")

    # 2. 找基准 Excel
    base_path, exists_today = find_base_excel()
    if base_path is None:
        print("错误：找不到基准 Excel，无法合并。")
        return

    print(f"\n基准文件: {base_path} {'(当天已存在)' if exists_today else '(从昨天复制)'}")

    if not exists_today:
        shutil.copy(base_path, XLSX_FILE)
        print(f"已创建: {XLSX_FILE}")

    # 3. 打开 Excel
    wb = openpyxl.load_workbook(XLSX_FILE)
    ws = wb[wb.sheetnames[0]]

    # 3.5 去重：如果当天已存在行，先删掉
    removed = remove_today_rows(ws, TODAY)
    if removed:
        print(f"已清理当天 {removed} 行旧数据，重新追加")

    # 4. 获取最后任务状态
    last_g, remaining = get_last_task_info(ws)
    print(f"上一个任务 G={last_g.strftime('%Y-%m-%d')}, 当天剩余 {remaining}h")

    # 5. 找最后一条数据行，紧贴其后插入
    last_data_row = 0
    for row in range(1, ws.max_row + 1):
        a_val = ws.cell(row=row, column=1).value
        if a_val and '填充说明' in str(a_val):
            break
        c_val = ws.cell(row=row, column=3).value
        if c_val is not None and str(c_val).strip().isdigit():
            last_data_row = row

    if last_data_row == 0:
        insert_pos = 2  # 空模板从第2行开始
        ref_row = 1     # 用表头做样式参考
    else:
        insert_pos = last_data_row + 1
        ref_row = last_data_row
    print(f"最后数据行: {last_data_row}, 插入起始行: {insert_pos}")

    # 6. 获取最大序号
    max_seq = 0
    for row in range(1, insert_pos):
        c_val = ws.cell(row=row, column=3).value
        if c_val is not None and str(c_val).strip().isdigit():
            max_seq = max(max_seq, int(str(c_val).strip()))

    # 7. 优先写已有空行，不够再在备注前插行（保留数据与备注之间的间隔）
    current_g = last_g
    current_remaining = remaining
    prev_repo = None

    # 找到上一个任务的 repo（从最后数据行往前找）
    for row in range(last_data_row, 0, -1):
        b_val = ws.cell(row=row, column=2).value
        if b_val:
            prev_repo = str(b_val).strip()
            break

    notes_row = find_notes_row(ws)
    gap_rows = notes_row - last_data_row - 1  # 数据与备注之间的空行数

    if len(tasks) > gap_rows:
        need = len(tasks) - gap_rows
        for _ in range(need):
            ws.insert_rows(notes_row)
        print(f"空行不足，在备注前补插 {need} 行")

    for i, task in enumerate(tasks):
        ws.insert_rows(insert_pos)

        # A 列：日期
        date_cell = ws.cell(row=insert_pos, column=1)
        if i == 0:
            date_cell.value = TODAY
            date_cell.number_format = 'yyyy/m/d;@'

        # B 列：项目名称（同项目省略）
        repo = task['repo']
        display_name = REPO_MAP.get(repo, repo)
        if repo != prev_repo:
            ws.cell(row=insert_pos, column=2).value = display_name
            prev_repo = repo

        # C 列：序号
        max_seq += 1
        ws.cell(row=insert_pos, column=3).value = max_seq

        # D 列：任务描述，按内容长度自适应行高
        ws.cell(row=insert_pos, column=4).value = task['desc']
        # 估算行高：中文约 35 字/行，每行 15pt，最小 30pt
        lines = max(1, len(task['desc']) / 35)
        ws.row_dimensions[insert_pos].height = max(30, lines * 15)

        # E 列：完成百分比（来自 md）
        ws.cell(row=insert_pos, column=5).value = task['pct']

        # F 列：任务创建时间
        f_cell = ws.cell(row=insert_pos, column=6)
        f_cell.value = TODAY
        f_cell.number_format = 'yyyy/m/d;@'

        # G 列：预计完成时间（按工作日叠加）
        current_g, current_remaining = calc_g_column(task['human_h'], current_g, current_remaining)
        g_cell = ws.cell(row=insert_pos, column=7)
        g_cell.value = current_g
        g_cell.number_format = 'yyyy/m/d;@'

        # H 列：人工工时
        ws.cell(row=insert_pos, column=8).value = task['human_h']

        # J 列：实际完成时间
        j_cell = ws.cell(row=insert_pos, column=10)
        j_cell.value = TODAY
        j_cell.number_format = 'yyyy/m/d;@'

        # K 列：AI 工时
        ws.cell(row=insert_pos, column=11).value = task['ai_h']

        # N 列：备注
        ws.cell(row=insert_pos, column=14).value = task['note']

        # 复制样式
        for col in range(1, 15):
            copy_style(ws.cell(row=ref_row, column=col), ws.cell(row=insert_pos, column=col))

        # D 列、N 列自动换行
        for col in (4, 14):
            ws.cell(row=insert_pos, column=col).alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

        print(f"  新增行 {insert_pos}: [{repo}] #{max_seq} G={current_g.strftime('%m-%d')} H={task['human_h']}h K={task['ai_h']}h")
        insert_pos += 1

    # 8. A 列合并当天所有行，居中
    start_row = insert_pos - len(tasks)
    end_row = insert_pos - 1
    if len(tasks) > 1:
        ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
        print(f"A 列合并: 行 {start_row}-{end_row}")
    for r in range(start_row, end_row + 1):
        ws.cell(row=r, column=1).alignment = Alignment(horizontal='center', vertical='center')

    # 8.5 确保数据与备注之间至少空 2 行
    notes_now = find_notes_row(ws)
    gap = notes_now - insert_pos
    if gap < 2:
        need = 2 - gap
        for _ in range(need):
            ws.insert_rows(notes_now)
        print(f"补空行: 数据行{insert_pos-1}与备注之间补 {need} 行，间距≥2")

    # 9. 调整列宽（与原表一致）、冻结表头，保存
    col_widths = {'A':13,'B':22.5,'C':14.8,'D':59.3,'E':16.4,'F':16.4,'G':15.9,'H':13.9,'I':24.3,'J':13,'K':14.8,'L':12,'M':41.8,'N':28.5}
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width
    ws.freeze_panes = 'A2'
    wb.save(XLSX_FILE)
    print(f"\n已保存: {XLSX_FILE}")
    print(f"新增 {len(tasks)} 行，序号 {max_seq - len(tasks) + 1}-{max_seq}")

if __name__ == '__main__':
    main()

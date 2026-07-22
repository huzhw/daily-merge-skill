# -*- coding: utf-8 -*-
"""每日脚本：找最新 Excel → 复制上工作日未完成任务 → 追加当天 md 新需求（按仓库分组合并）"""
import openpyxl, os, shutil, re
from datetime import datetime, timedelta
from copy import copy
from openpyxl.styles import Alignment

DESKTOP = r"C:\Users\Administrator\Desktop"
TODAY = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
YEAR = TODAY.strftime("%Y")
MONTH = TODAY.strftime("%m")
MM = TODAY.strftime("%m")
DD = TODAY.strftime("%d")

REPO_MAP = {
    'lanxum-amisp': '档案V6', 'lanxum-amisp-java': '档案V6', 'lanxum-amisp-react': '档案V6',
    'workingpaper-v5.5': '中信底稿V5',
    'standard_thdg_zxdm': '中信底稿V5',
    # md 直接用中文名的情况
    '中信底稿V5': '中信底稿V5',
    '档案V6': '档案V6',
    '档案系统V6': '档案V6',
    '智能数据底座': '智能数据底座',
}

REPORT_DIR = os.path.join(DESKTOP, f"报告-{YEAR}年", f"日报-{YEAR}-{MONTH}月")
MD_FILE = os.path.join(REPORT_DIR, f"日报需求记录-{YEAR}-{MM}-{DD}.md")
XLSX_FILE = os.path.join(REPORT_DIR, f"日报表格-胡志伟~~{MM}-{DD}.xlsx")


def find_latest_xlsx():
    """找目录下日期最新且 < TODAY 的 Excel 文件路径，兼容周末/节假日"""
    if not os.path.exists(REPORT_DIR):
        return None
    pattern = re.compile(r'日报表格-胡志伟~~(\d{2})-(\d{2})\.xlsx$')
    best_path = None
    best_date = None
    for fname in os.listdir(REPORT_DIR):
        m = pattern.match(fname)
        if not m:
            continue
        f_mm, f_dd = m.groups()
        try:
            f_date = datetime(TODAY.year, int(f_mm), int(f_dd))
        except ValueError:
            continue
        if f_date < TODAY and (best_date is None or f_date > best_date):
            best_date = f_date
            best_path = os.path.join(REPORT_DIR, fname)
    return best_path


def parse_date(val):
    """统一解析单元格日期值 → date 对象"""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    try:
        return datetime.strptime(str(val)[:10], '%Y-%m-%d').date()
    except:
        return None


def parse_pct(val):
    """解析 E 列完成百分比 → 0~1 浮点数，非 100% 即未完成"""
    if val is None:
        return 0
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).replace('%', '').strip()
    try:
        return float(s) / 100.0 if '%' in str(val) or float(s) <= 1 else float(s) / 100.0
    except:
        return 0


def parse_md(filepath):
    """解析 md 文件，按仓库分组合并已完成任务。

    返回 list[dict]: [{repo, desc(合并:需求概述+涉及模块+备注), human_h(sum)}, ...]
    """
    if not os.path.exists(filepath):
        raise FileNotFoundError(f"md 文件不存在: {filepath}")
    with open(filepath, 'r', encoding='utf-8') as f:
        content = f.read()

    tasks_by_repo = {}
    repo_order = []
    in_table = False

    for line in content.split('\n'):
        line = line.strip()
        if line.startswith('| 序号 |'):
            in_table = True
            continue
        if line.startswith('|------|'):
            continue
        if not in_table or not line.startswith('|'):
            continue
        if line.startswith('| >') or '空行为模板' in line:
            break

        cells = [c.strip() for c in line.split('|')[1:-1]]
        if len(cells) < 9:
            continue
        seq, date, repo, desc, modules, status, human_h, ai_h, note = cells[:9]

        is_done = status in ('已完成', '100%')
        if not is_done or not seq.isdigit():
            continue

        display = REPO_MAP.get(repo, repo)
        if display not in tasks_by_repo:
            tasks_by_repo[display] = []
            repo_order.append(display)

        tasks_by_repo[display].append({
            'desc': desc,
            'modules': modules,
            'note': note,
            'human_h': float(human_h),
            'ai_h': float(ai_h),
        })

    result = []
    for repo in repo_order:
        tasks = tasks_by_repo[repo]
        desc_parts = []
        for i, t in enumerate(tasks, 1):
            part = f"{i}、{t['desc']}"
            if t['modules'] and t['modules'] != '—':
                part += f"\n{t['modules']}"
            if t['note'] and t['note'] != '—':
                part += f"\n{t['note']}"
            desc_parts.append(part)

        result.append({
            'repo': repo,
            'desc': '\n\n'.join(desc_parts),
            'human_h': sum(t['human_h'] for t in tasks),
        })

    return result


def next_workday(d):
    d = d + timedelta(days=1)
    while d.weekday() >= 5:
        d = d + timedelta(days=1)
    return d


def get_last_info(ws):
    """读最后一个数据行的 G 列和当天累计工时"""
    last_g = TODAY
    remaining = 0
    last_row = 0
    for row in range(1, ws.max_row + 1):
        a = ws.cell(row=row, column=1).value
        if a and '填充说明' in str(a):
            break
        c = ws.cell(row=row, column=3).value
        if c is not None and str(c).strip().isdigit():
            last_row = row
    if last_row > 0:
        gv = ws.cell(row=last_row, column=7).value
        if gv:
            if isinstance(gv, datetime):
                last_g = gv
            else:
                try:
                    last_g = datetime.strptime(str(gv)[:10], '%Y-%m-%d')
                except:
                    pass
        day_hours = 0
        for row in range(last_row, 0, -1):
            rd = parse_date(ws.cell(row=row, column=7).value)
            if rd is None or rd != last_g.date():
                break
            hv = ws.cell(row=row, column=8).value
            if hv:
                try:
                    day_hours += float(hv)
                except:
                    pass
        remaining = day_hours % 8
    return last_g, remaining, last_row


def find_notes(ws):
    """找「填充说明」起始行"""
    for row in range(1, ws.max_row + 1):
        if ws.cell(row=row, column=1).value and '填充说明' in str(ws.cell(row=row, column=1).value):
            return row
    return ws.max_row + 3


def calc_g(hours, prev_date, remaining):
    """8h/工作日叠加，跳过周六日"""
    total = remaining + hours
    d = prev_date
    while total > 8:
        total -= 8
        d = next_workday(d)
    return d, total


def copy_style(src, dst):
    """复制字号字体（不含颜色），+ 边框 + 对齐"""
    if src.has_style:
        from openpyxl.styles import Font
        sf = src.font
        dst.font = Font(name=sf.name, size=sf.size, bold=sf.bold, italic=sf.italic)
        dst.border = copy(src.border)
        dst.alignment = copy(src.alignment)




def find_previous_workday_rows(ws, notes_row):
    """找上一个工作日日期块中所有未完成行（E < 1），返回 [(row_num, date_val), ...]"""
    # 找到所有日期块及其行范围
    date_blocks = []  # [(date, start_row, end_row), ...]
    i = 2
    while i < notes_row:
        av = ws.cell(row=i, column=1).value
        d = parse_date(av)
        if d is not None and d < TODAY.date():
            # 找这个日期块的结束行（下一个非空 A 列或 notes 之前）
            j = i
            while j + 1 < notes_row:
                next_a = ws.cell(row=j + 1, column=1).value
                if next_a is not None:
                    break
                j += 1
            date_blocks.append((d, i, j))
            i = j + 1
        else:
            i += 1

    if not date_blocks:
        return []

    # 取最后一个日期块（上一个工作日）
    last_block = date_blocks[-1]
    prev_date, block_start, block_end = last_block

    # 扫描该日期块内 E < 1 的行（必须有序号才算数据行，跳过空行）
    unfinished = []
    for row in range(block_start, block_end + 1):
        cv = ws.cell(row=row, column=3).value
        if cv is None or not str(cv).strip().isdigit():
            continue  # 跳过空行、无序号行
        e_val = ws.cell(row=row, column=5).value
        pct = parse_pct(e_val)
        if pct < 1:
            unfinished.append(row)

    return unfinished


def remove_today_rows(ws):
    """清理当天已存在的旧数据行，支持去重"""
    today_merged = set()
    for mr in list(ws.merged_cells.ranges):
        if mr.min_col == 1 and mr.max_col == 1:
            d = parse_date(ws.cell(row=mr.min_row, column=1).value)
            if d == TODAY.date():
                for r in range(mr.min_row, mr.max_row + 1):
                    today_merged.add(r)
                try:
                    ws.unmerge_cells(str(mr))
                except:
                    pass
    rows_to_del = set()
    notes = find_notes(ws)
    for row in range(1, notes):
        d = parse_date(ws.cell(row=row, column=1).value)
        if d == TODAY.date():
            rows_to_del.add(row)
    rows_to_del |= today_merged
    for row in sorted(rows_to_del, reverse=True):
        ws.delete_rows(row)
    return len(rows_to_del)


def main():
    # ── 1. 预解析 md（只调一次） ──
    new_tasks = []
    if os.path.exists(MD_FILE):
        new_tasks = parse_md(MD_FILE)
        if new_tasks:
            print(f"md 新任务: {len(new_tasks)} 个仓库分组")
        else:
            print("md 无已完成任务")
    else:
        print("md 文件不存在，仅处理跨天复制")

    # ── 2. 找最新 Excel ──
    latest_xlsx = find_latest_xlsx()
    if not latest_xlsx:
        print(f"错误：找不到本月已有的 Excel 文件")
        return
    print(f"基准文件: {os.path.basename(latest_xlsx)}")
    shutil.copy(latest_xlsx, XLSX_FILE)

    wb = openpyxl.load_workbook(XLSX_FILE)
    ws = wb[wb.sheetnames[0]]

    # ── 3. 去重 ──
    removed = remove_today_rows(ws)
    if removed:
        print(f"已清理当天 {removed} 行旧数据")

    notes_row = find_notes(ws)

    # ── 4. 读取上工作日未完成任务数据（在 insert 前读完，避免行偏移） ──
    unfinished_src_rows = find_previous_workday_rows(ws, notes_row)
    print(f"上工作日未完成任务: {len(unfinished_src_rows)} 行")

    COPY_COLS = [2, 3, 4, 5, 6, 7, 8, 14]  # B/C/D/E/F/G/H/N
    unfinished_data = []
    for src_row in unfinished_src_rows:
        row_data = {}
        for col in COPY_COLS:
            src_cell = ws.cell(row=src_row, column=col)
            row_data[col] = {
                'value': src_cell.value,
                'number_format': src_cell.number_format,
            }
        unfinished_data.append(row_data)

    # ── 5. 计算插入位置 + 补空行 ──
    last_g, remaining, last_data_row = get_last_info(ws)
    insert_pos = last_data_row + 1

    total_new = len(unfinished_data) + len(new_tasks)
    if total_new == 0:
        print("无未完成任务、无新任务，跳过。")
        return

    gap = notes_row - last_data_row - 1
    if total_new > gap:
        need = total_new - gap
        for _ in range(need):
            ws.insert_rows(notes_row)
        print(f"空行不足，补插 {need} 行")

    # ── 6. 写入未完成任务 ──
    for row_data in unfinished_data:
        ws.insert_rows(insert_pos)

        # A 列：当天日期
        ws.cell(row=insert_pos, column=1).value = TODAY
        ws.cell(row=insert_pos, column=1).number_format = 'yyyy/m/d;@'

        for col, data in row_data.items():
            dst_cell = ws.cell(row=insert_pos, column=col)
            # E 列统一写 0%，其余列保留原值
            if col == 5:
                dst_cell.value = '0%'
            else:
                dst_cell.value = data['value']
            dst_cell.number_format = data['number_format']

        ws.row_dimensions[insert_pos].height = 35

        # 所有列统一从表头复制基本样式（字体/边框/对齐，不含颜色）
        for col in range(1, 15):
            copy_style(ws.cell(row=1, column=col), ws.cell(row=insert_pos, column=col))
        # D/M/N 列左对齐
        ws.cell(row=insert_pos, column=4).alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        for col in (13, 14):
            ws.cell(row=insert_pos, column=col).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

        print(f"  复制未完成行 → {insert_pos}: C={ws.cell(row=insert_pos, column=3).value} E=0%")
        insert_pos += 1

    # ── 7. 追加 md 新任务（按仓库合并） ──
    if new_tasks:
        current_g = last_g
        current_remaining = remaining
        prev_repo = None
        max_seq = 0

        for row in range(1, insert_pos):
            cv = ws.cell(row=row, column=3).value
            if cv is not None and str(cv).strip().isdigit():
                max_seq = max(max_seq, int(str(cv).strip()))

        for row in range(last_data_row, 0, -1):
            bv = ws.cell(row=row, column=2).value
            if bv:
                prev_repo = str(bv).strip()
                break

        for task in new_tasks:
            ws.insert_rows(insert_pos)

            ws.cell(row=insert_pos, column=1).value = TODAY
            ws.cell(row=insert_pos, column=1).number_format = 'yyyy/m/d;@'

            repo = task['repo']
            if repo != prev_repo:
                ws.cell(row=insert_pos, column=2).value = repo
                prev_repo = repo

            max_seq += 1
            ws.cell(row=insert_pos, column=3).value = max_seq

            ws.cell(row=insert_pos, column=4).value = task['desc']

            ws.cell(row=insert_pos, column=5).value = '0%'

            cell_f = ws.cell(row=insert_pos, column=6)
            cell_f.value = TODAY
            cell_f.number_format = 'yyyy/m/d;@'

            current_g, current_remaining = calc_g(task['human_h'], current_g, current_remaining)
            g_cell = ws.cell(row=insert_pos, column=7)
            g_cell.value = current_g
            g_cell.number_format = 'yyyy/m/d;@'

            ws.cell(row=insert_pos, column=8).value = task['human_h']

            for c in range(1, 15):
                copy_style(ws.cell(row=1, column=c), ws.cell(row=insert_pos, column=c))
            # D/M/N 列左对齐
            ws.cell(row=insert_pos, column=4).alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            for col in (13, 14):
                ws.cell(row=insert_pos, column=col).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            ws.row_dimensions[insert_pos].height = 35

            print(f"  新增行 {insert_pos}: [{repo}] #{max_seq} G={current_g.strftime('%m-%d')} H={task['human_h']}h")
            insert_pos += 1

    # ── 8. A 列合并当天所有行 ──
    today_start = None
    today_end = None
    for row in range(2, insert_pos):
        d = parse_date(ws.cell(row=row, column=1).value)
        if d == TODAY.date():
            if today_start is None:
                today_start = row
            today_end = row
    if today_start is not None and today_end is not None and today_end > today_start:
        ws.merge_cells(start_row=today_start, start_column=1, end_row=today_end, end_column=1)
    if today_start is not None:
        for r in range(today_start, (today_end or today_start) + 1):
            ws.cell(row=r, column=1).alignment = Alignment(horizontal='center', vertical='center')

    # ── 9. 间距、列宽、冻结 ──
    notes_now = find_notes(ws)
    g = notes_now - insert_pos
    if g < 2:
        need = 2 - g
        for _ in range(need):
            ws.insert_rows(notes_now)
        print(f"补空行 {need} 行")

    cw = {'A': 13, 'B': 23, 'C': 15, 'D': 60, 'E': 16, 'F': 18, 'G': 18, 'H': 14,
          'I': 32, 'J': 13, 'K': 15, 'L': 14, 'M': 25, 'N': 60}
    for k, v in cw.items():
        ws.column_dimensions[k].width = v
    ws.freeze_panes = 'A2'
    ws.sheet_view.topLeftCell = 'A1'

    wb.save(XLSX_FILE)
    print(f"\n已保存: {XLSX_FILE}")


if __name__ == '__main__':
    main()

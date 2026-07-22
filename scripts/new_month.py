# -*- coding: utf-8 -*-
"""月初脚本：从模板起新表，写入当天已完成需求（按仓库分组合并）"""
import openpyxl, os
from datetime import datetime, timedelta
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

DESKTOP = r"C:\Users\Administrator\Desktop"
TODAY = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
Y = TODAY.strftime("%Y"); M = TODAY.strftime("%m"); MM = TODAY.strftime("%m"); DD = TODAY.strftime("%d")

REPO_MAP = {
    'lanxum-amisp': '档案V6', 'lanxum-amisp-java': '档案V6', 'lanxum-amisp-react': '档案V6',
    'workingpaper-v5.5': '中信底稿V5',
    'standard_thdg_zxdm': '中信底稿V5',
    # md 直接用中文名的情况
    '中信底稿V5': '中信底稿V5',
    '档案V6': '档案V6',
    '档案系统V6': '档案V6',
    '智能数据底座': '智能数据底座',
}

RD = os.path.join(DESKTOP, f"报告-{Y}年", f"日报-{Y}-{M}月")
MD = os.path.join(RD, f"日报需求记录-{Y}-{MM}-{DD}.md")
XL = os.path.join(RD, f"日报表格-胡志伟~~{MM}-{DD}.xlsx")


def nwd(d):
    d += timedelta(days=1)
    while d.weekday() >= 5:
        d += timedelta(days=1)
    return d


def parse():
    """解析 md，按仓库分组合并已完成任务"""
    ts_by_repo = {}
    repo_order = []
    with open(MD, encoding='utf-8') as f:
        c = f.read()
    for l in c.split('\n'):
        l = l.strip()
        if not l.startswith('|') or '---' in l or '序号' in l or '空行' in l:
            continue
        p = [x.strip() for x in l.split('|')[1:-1]]
        if len(p) < 9:
            continue
        seq, date, repo, desc, modules, status, human_h, ai_h, note = p[:9]
        if status not in ('已完成', '100%') or not seq.isdigit():
            continue

        display = REPO_MAP.get(repo, repo)
        if display not in ts_by_repo:
            ts_by_repo[display] = []
            repo_order.append(display)
        ts_by_repo[display].append({
            'desc': desc,
            'modules': modules,
            'note': note,
            'human_h': float(human_h),
            'ai_h': float(ai_h),
        })

    result = []
    for repo in repo_order:
        tasks = ts_by_repo[repo]
        desc_parts = []
        for i, t in enumerate(tasks, 1):
            part = f"{i}、{t['desc']}"
            if t['modules'] and t['modules'] != '—':
                part += f"\n{t['modules']}"
            if t['note'] and t['note'] != '—':
                part += f"\n{t['note']}"
            desc_parts.append(part)

        result.append({
            'repo': repo,
            'desc': '\n\n'.join(desc_parts),
            'human_h': sum(t['human_h'] for t in tasks),
        })

    return result


def main():
    ts = parse()
    if not ts:
        print('无任务')
        return
    os.makedirs(RD, exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '日报'

    # 样式
    hfont = Font(name='Microsoft YaHei', size=11, bold=True, color='FFFFFF')
    hfill = PatternFill(start_color='1450B8', end_color='1450B8', fill_type='solid')
    halign = Alignment(horizontal='center', vertical='center', wrap_text=True)
    dfont = Font(name='Microsoft YaHei', size=11)
    dalign = Alignment(horizontal='center', vertical='center')
    bthin = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
    bdata = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))

    # 表头
    hd = ['*日期', '*项目名称', '*序号\n（全局唯一）', '*任务描述', '*完成百分比',
          '*任务创建时间\n（录入后勿修改）', '*预计完成时间\n（录入后勿修改）',
          '*预计工作人时\n', '调整预计完成时间\n（由于优先级问题调整时填写）',
          '实际完成时间', '*实际工作人时', '插队序号\n若有插队必填', '延期/调整原因', '备注/说明']
    for i, h in enumerate(hd, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = hfont; c.fill = hfill; c.alignment = halign; c.border = bthin
    ws.row_dimensions[1].height = 57

    # 备注
    notes = [
        ('填充说明：\n填写时例子请删除', 1),
        ('1、带*号的字段必填。', 2),
        ('2、前一天已完成的任务不需要重新出现在当天的记录中（若之前的任务完成后有bug需要修改，需要新建一条记录，复制之前任务的描述后面增加【bug】字样，并在插队序号中填写之前的任务序号，主序号依然递增）', 2),
        ('3、如果遇到更优先级的任务阻断了当前任务，需要新增阻断任务的信息，在插队序号中写入被影响的任务序号，多个任务逗号隔开，并插入/修改被影响任务的【调整预计完成时间】字段', 2),
        ('4、工时采用小时单位，按一天8小时计算，其中实际工时每天更新一直到任务完成100%，任务完成的当天做最后一次记录更新后，第二天不用重复写入该记录。', 2),
        ('5、此表在一个自然月内持续迭代，下一个自然月后重新起一个新表。', 2),
        ('6、备注按需填写，延迟原因需要简要写明。', 2),
        ('7、每天日报发送到公邮：kpi-daily-report@lscjz.com；每人每天一封。', 2),
        ('8、开会和沟通也当作任务记录时长', 2),
    ]
    nr = len(ts) + 5
    for i, (txt, col) in enumerate(notes):
        ws.cell(row=nr + i, column=col, value=txt)

    # 写数据
    row = 2
    gd = TODAY
    gr = 0
    repo_prev = None
    seq = 0

    for t in ts:
        # A 列：当天日期
        ws.cell(row=row, column=1, value=TODAY).number_format = 'yyyy/m/d;@'
        ws.cell(row=row, column=1).font = dfont; ws.cell(row=row, column=1).alignment = dalign; ws.cell(row=row, column=1).border = bdata

        # B 列：项目名（同项目省略）
        if t['repo'] != repo_prev:
            ws.cell(row=row, column=2, value=t['repo'])
            repo_prev = t['repo']
        ws.cell(row=row, column=2).font = dfont; ws.cell(row=row, column=2).alignment = dalign; ws.cell(row=row, column=2).border = bdata

        # C 列：序号
        seq += 1
        ws.cell(row=row, column=3, value=seq)
        ws.cell(row=row, column=3).font = dfont; ws.cell(row=row, column=3).alignment = dalign; ws.cell(row=row, column=3).border = bdata

        # D 列：合并后的描述，行高固定 35
        ws.cell(row=row, column=4, value=t['desc'])
        ws.cell(row=row, column=4).font = dfont
        ws.cell(row=row, column=4).alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        ws.cell(row=row, column=4).border = bdata
        ws.row_dimensions[row].height = 35

        # E 列：固定 0%
        ws.cell(row=row, column=5, value='0%')
        ws.cell(row=row, column=5).font = dfont; ws.cell(row=row, column=5).alignment = dalign; ws.cell(row=row, column=5).border = bdata

        # F 列：任务创建时间 = 当天
        ws.cell(row=row, column=6, value=TODAY).number_format = 'yyyy/m/d;@'
        ws.cell(row=row, column=6).font = dfont; ws.cell(row=row, column=6).alignment = dalign; ws.cell(row=row, column=6).border = bdata

        # G 列：按总工时叠加计算
        gr2 = gr + t['human_h']
        gd2 = gd
        while gr2 > 8:
            gr2 -= 8
            gd2 = nwd(gd2)
        ws.cell(row=row, column=7, value=gd2).number_format = 'yyyy/m/d;@'
        ws.cell(row=row, column=7).font = dfont; ws.cell(row=row, column=7).alignment = dalign; ws.cell(row=row, column=7).border = bdata
        gd = gd2; gr = gr2

        # H 列：总工时
        ws.cell(row=row, column=8, value=t['human_h'])
        ws.cell(row=row, column=8).font = dfont; ws.cell(row=row, column=8).alignment = dalign; ws.cell(row=row, column=8).border = bdata

        # I/J/K 列：留空，只加边框
        for c in (9, 10, 11, 12, 13, 14):
            ws.cell(row=row, column=c).border = bdata

        print(f'Row{row}: {t["repo"]} #{seq} G={gd2.strftime("%m-%d")} H={t["human_h"]}h')
        row += 1

    # A 列合并
    if len(ts) > 1:
        ws.merge_cells(start_row=2, start_column=1, end_row=row - 1, end_column=1)

    # 列宽
    cw = {'A': 13, 'B': 23, 'C': 15, 'D': 60, 'E': 16, 'F': 18, 'G': 18, 'H': 14,
          'I': 32, 'J': 13, 'K': 15, 'L': 14, 'M': 25, 'N': 60}
    for k, v in cw.items():
        ws.column_dimensions[k].width = v
    ws.freeze_panes = 'A2'
    ws.sheet_view.topLeftCell = 'A1'
    wb.save(XL)
    print(f'Done: {XL}')


if __name__ == '__main__':
    main()

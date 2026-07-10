# -*- coding: utf-8 -*-
"""生成空白日报模板，用于每月初新表"""
import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = '日报'

# === 表头 ===
headers = [
    '*日期', '*项目名称', '*序号\n（全局唯一）', '*任务描述', '*完成百分比',
    '*任务创建时间\n（录入后勿修改）', '*预计完成时间\n（录入后勿修改）',
    '*预计工作人时\n', '调整预计完成时间\n（由于优先级问题调整时填写）',
    '实际完成时间', '*实际工作人时', '插队序号\n若有插队必填',
    '延期/调整原因', '备注/说明'
]

header_font = Font(bold=True, size=10)
header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.alignment = header_align

ws.row_dimensions[1].height = 45

# === 空数据行 ===
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
for r in range(2, 6):
    for c in range(1, 15):
        cell = ws.cell(row=r, column=c)
        cell.border = thin_border
        cell.alignment = Alignment(vertical='center', wrap_text=True)
    ws.row_dimensions[r].height = 25

# === 备注说明 ===
notes = [
    '填充说明：\n填写时例子请删除',
    '1、带*号的字段必填。',
    '2、前一天已完成的任务不需要重新出现在当天的记录中（若之前的任务完成后有bug需要修改，需要新建一条记录，复制之前任务的描述后面增加【bug】字样，并在插队序号中填写之前的任务序号，主序号依然递增）',
    '3、如果遇到更优先级的任务阻断了当前任务，需要新增阻断任务的信息，在插队序号中写入被影响的任务序号，多个任务逗号隔开，并插入/修改被影响任务的【调整预计完成时间】字段',
    '4、工时采用小时单位，按一天8小时计算，其中实际工时每天更新一直到任务完成100%，任务完成的当天做最后一次记录更新后，第二天不用重复写入该记录。',
    '5、此表在一个自然月内持续迭代，下一个自然月后重新起一个新表。',
    '6、备注按需填写，延迟原因需要简要写明。',
    '7、每天日报发送到公邮：kpi-daily-report@lscjz.com；每人每天一封。',
    '8、开会和沟通也当作任务记录时长',
]

note_start = 9  # 和数据之间空了几行
for i, note in enumerate(notes):
    row = note_start + i
    if i == 0:
        ws.cell(row=row, column=1).value = note
    else:
        ws.cell(row=row, column=2).value = note

# === 列宽 ===
col_widths = {
    'A': 14, 'B': 18, 'C': 10, 'D': 60, 'E': 12,
    'F': 18, 'G': 18, 'H': 14, 'I': 18, 'J': 14,
    'K': 14, 'L': 14, 'M': 16, 'N': 60
}
for col_letter, width in col_widths.items():
    ws.column_dimensions[col_letter].width = width

# === 冻结表头 ===
ws.freeze_panes = 'A2'

import os
out_dir = os.path.dirname(os.path.abspath(__file__))
out_path = os.path.join(out_dir, '日报模板.xlsx')
wb.save(out_path)
print(f'已生成: {out_path}')
